import csv
import json
import os
import re
import urllib.request
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from Log.logs import Logs
from openpyxl import Workbook
from robocorp import storage, workitems
from helpers.article import Article
from helpers.payload import Payload
from helpers.selector import Selector
from webdriver_util.webdrv_util import *
from selenium.webdriver.chrome.options import Options
from urllib.parse import unquote

load_dotenv("config/.env")


class ProducerMethods:
    @staticmethod
    def read_csv_create_work_item(debug: bool = False):
        """
        Reads a CSV file and creates work items from its data.

        Args:
            debug (bool): If True, the function returns the payload instead of creating work items.

        Returns:
            Payload | None: Returns payload if debug is True, otherwise None.
        """
        csv_file_path = os.path.join("devdata", "csv_input.csv")
        if os.path.exists(csv_file_path):
            with open(csv_file_path, mode="r", newline="") as file:
                reader = csv.reader(file)
                header = next(reader)
                for row in reader:
                    payload = Payload(
                        phrase_test=row[0],
                        section=row[1],
                        data_range=int(row[2]),
                        sort_by=int(row[3]),
                        results=int(row[4]),
                    )
                    if not debug:
                        workitems.outputs.create(
                            payload={
                                "phrase_test": payload.phrase_test,
                                "section": payload.section,
                                "data_range": payload.data_range,
                                "sort_by": payload.sort_by,
                                "results": payload.results,
                            }
                        )
                    else:
                        return payload
        else:
            logger.critical(f"The CSV file: {csv_file_path} was not found.")


class ScraperMethods:
    @staticmethod
    def get_work_item() -> Payload | None:
        """
        Retrieves the current work item and converts it to a Payload object.

        Returns:
            Payload | None: Returns the payload if a work item exists, otherwise None.
        """
        item = workitems.inputs.current
        if item:
            logger.info("Received payload:", item.payload)
            pay = Payload(
                phrase_test=item.payload["phrase_test"],
                section=item.payload["section"],
                data_range=int(item.payload["data_range"]),
                sort_by=int(item.payload["sort_by"]),
                results=int(item.payload["results"]),
            )
            return pay
        else:
            logger.critical("An error occurred during the process!")

    @staticmethod
    def inicial_search(driver: Selenium, phrase: str):
        """
        Performs the initial search on the website.

        Args:
            driver (Selenium): The Selenium driver instance.
            phrase (str): The search phrase.

        Returns:
            bool: True if the search was initiated successfully, otherwise False.
        """
        try:
            logger.info("Starting Scraper")
            search = find_element(
                driver.driver, Selector(css='button[data-element="search-button"]')
            )
            if search:
                center_element(driver.driver, search)
                click_elm(driver.driver, search)
                search_field = find_element(
                    driver.driver,
                    Selector(css="input[data-element='search-form-input']"),
                )
                if search_field:
                    center_element(driver.driver, search_field)
                    slow_send_keys(search_field, phrase + Keys.ENTER, False)
                    return True
        except Exception as e:
            print(e.with_traceback())
            return False

    @staticmethod
    def fine_search(
        driver: WebDriver,
        phrase: str,
        section: str,
        sort_by: int = 0,
        data_range: int = 0,
    ):
        """
        Performs fine-tuned search with additional filters.

        Args:
            driver (WebDriver): The WebDriver instance.
            phrase (str): The search phrase.
            section (str): The section to filter.
            sort_by (int): The sort option (default is 0).
            data_range (int): The data range option (default is 0).

        Returns:
            tuple: (bool, int) indicating success and the data range used.
        """
        try:
            no_results_match = find_element(
                driver.driver,
                Selector(css="div[class='search-results-module-no-results']"),
            )
            if no_results_match:
                logger.critical("No search match found.")
                return False, 0

            # Expand Filter
            label_search = find_element(
                driver.driver, Selector(css="span[class='see-all-text']")
            )
            if label_search:
                center_element(driver.driver, label_search)
                click_elm(driver.driver, label_search)
                sleep(1.5)
                wait_for_modal(driver.driver)
                if len(section.strip()) > 0:
                    list_topics = extract_names_from_list_items(driver)
                    if list_topics:
                        element_topic, topic = search_and_click_topics(
                            driver.driver, list_topics, section
                        )
                        if element_topic == False and topic == False:
                            return False, 0

                if sort_by > 0:  # not Relevance (default)
                    sleep(1.5)
                    select_sort_by = find_element(
                        driver.driver, Selector(css="select[name='s']")
                    )
                    if select_sort_by:
                        if sort_by in [1, 2]:
                            sleep(1.5)
                            center_element(driver.driver, select_sort_by)
                            sleep(0.5)
                            select_option_value(select_sort_by, sort_by)
                            sleep(1.5)
                        else:
                            logger.error(f"Sort parameter does not exist: {sort_by}")
                            logger.info("Relevance is selected")

                if data_range >= 0:
                    # LEGEND: 0= Actual Page, 1= Results you want, 2= All Results
                    data_range_str = ["Actual Page", "Results you want", "All Results"][
                        data_range
                    ]
                    logger.info(f"{data_range_str} is selected")
                return True, data_range
        except Exception as e:
            logger.critical(f"An error occurred: {e.__cause__}.")
            return False, 0

    @staticmethod
    def collect_articles(
        driver: WebDriver, data_range: int = 0
    ) -> list[Article] | None:
        """
        Collects articles from the search results.

        Args:
            driver (WebDriver): The WebDriver instance.
            data_range (int): The data range option.

        Returns:
            list[Article] | None: List of collected articles or None if an error occurs.
        """
        try:
            list_articles = []
            more_results = True
            cont = 1
            while more_results:
                logger.info("Search results found")
                search_results_section = find_element(
                    driver.driver,
                    Selector(css="div[class='search-results-module-results-header']"),
                )
                if search_results_section:
                    logger.info("Search results found")
                    wait_for_modal(driver.driver)
                    sleep(2.5)
                    li_search_results = find_all_css(
                        driver.driver,
                        'ul[class*="search-results-module-results-menu"] li',
                    )
                    if li_search_results:
                        sleep(3.5)
                        for li in li_search_results:
                            logger.info(f"Creating an article object: {cont}")
                            article = Article()
                            title = li.find_element(
                                By.CSS_SELECTOR, "h3[class='promo-title']"
                            )
                            time = li.find_element(
                                By.CSS_SELECTOR, "p[class^='promo-timestamp']"
                            )
                            description = li.find_element(
                                By.CSS_SELECTOR, "p[class='promo-description']"
                            )
                            try:
                                center_element(driver.driver, li)
                                photo = find_elm_picture(
                                    li, Selector(css='img[src*=".jpg"]')
                                )
                                if photo:
                                    article.picture_filename = photo
                                    logger.info(
                                        f"Picture found: {article.picture_filename}"
                                    )
                            except:
                                logger.info("Picture information in article not found.")
                                pass
                            logger.info("Article information found.")
                            article.title = title.text.strip()
                            article.description = description.text.strip()
                            time_str = time.text.strip()
                            parse = parse_time_ago(time_str)
                            if parse:
                                article.date = parse
                            else:
                                article.date = datetime.strptime(
                                    time.text.strip(), "%B %d, %Y"
                                )
                            logger.info(
                                f"Title: {article.title} -- Date: {article.date}"
                            )
                            sleep(0.4)
                            list_articles.append(article)
                            if data_range == cont:
                                more_results = False
                                break
                            cont += 1
                        button_next = find_element(
                            driver.driver,
                            Selector(
                                css='div[class="search-results-module-next-page"]'
                            ),
                        )
                        if button_next:
                            center_element(driver.driver, button_next)
                            click_elm(driver.driver, button_next)
                            if data_range < cont:
                                more_results = False
            return list_articles
        except Exception as e:
            logger.critical(f"An error occurred: {e.__cause__}.")
            return None


class ExcelOtherMethods:
    @staticmethod
    def __extract_filename_from_url(url):
        """
        Extracts the filename from a URL.

        Args:
            url (str): The URL of the file.

        Returns:
            str | None: The extracted filename or None if no match found.
        """
        match = re.search(r"[^/]+$", url)
        if match:
            filename_with_extension = match.group(0)
            filename_with_extension = unquote(filename_with_extension)
            filename_with_extension = re.sub(
                r'[<>:"/\\|?*]', "", filename_with_extension
            )
            return filename_with_extension
        else:
            return None

    @staticmethod
    def __contains_money(text):
        """
        Checks if the text contains monetary amounts.

        Args:
            text (str): The input text.

        Returns:
            bool: True if monetary amounts are found, otherwise False.
        """
        pattern = r"\$[0-9,.]+|\b\d+\s*(?:dollars|USD)\b"
        matches = re.findall(pattern, text)
        return bool(matches)

    @staticmethod
    def __download_image(url: str, article_title: str) -> str:  
        """  
        Downloads an image from a given URL with a filename based on the article title.  
        
        Args:  
            url (str): The URL of the image.  
            article_title (str): The title of the article.  
        
        Returns:  
            str: The local path where the image is saved.  
        """  
        # Extract the file extension from the URL  
        file_extension = os.path.splitext(url)[-1]  
        
        # Sanitize the article title and ensure it is no more than 20 characters  
        sanitized_title = re.sub(r'[<>:"/\\|?*]', '', article_title)  
        short_title = sanitized_title[:20]  
        
        # Create the full filename  
        filename = f"{short_title}{file_extension}"  
        
        # Define the download path  
        project_dir = str(os.getcwd())  
        full_path = Path(project_dir, "devdata", "downloads")  
        
        if os.path.isdir(full_path):  
            full_path = os.path.join(full_path, filename)  
            logger.info(f"Downloading image: {url}")  
            urllib.request.urlretrieve(url, full_path)  
            return full_path 

    @staticmethod
    def prepare_articles(list_articles: list[Article], phrase: str) -> list[Article]:
        """
        Prepares articles by adding additional metadata.

        Args:
            list_articles (list[Article]): List of articles to be prepared.
            phrase (str): The phrase to count in titles and descriptions.

        Returns:
            list[Article]: List of prepared articles.
        """
        new_list_articles = []
        if list_articles:
            for article in list_articles:
                art = Article()
                art.title = article.title
                art.date = article.date
                art.title_count_phrase = len(
                    re.findall(re.escape(phrase), article.title.strip(), re.IGNORECASE)
                )
                art.description = article.description
                art.description_count_phrase = len(
                    re.findall(re.escape(phrase), article.description.strip(), re.IGNORECASE)
                )
                art.find_money_title_description = ExcelOtherMethods.__contains_money(
                    article.title
                )
                if len(article.picture_filename) > 0:
                    art.picture_filename = article.picture_filename
                    art.picture_local_path = ExcelOtherMethods.__download_image(
                        art.picture_filename,
                        article.title.strip()
                    )
                new_list_articles.append(art)
                logger.info(f"Article created: {art.to_dict()}")
            return new_list_articles

    @staticmethod
    def export_excel(list_articles: list[Article]):
        """
        Exports the list of articles to an Excel file.

        Args:
            list_articles (list[Article]): List of articles to be exported.
        """
        project_dir = str(os.getcwd())
        full_path = Path(project_dir, "output")
        excel_file_path = os.path.join(full_path, "Articles.xlsx")
        wb = Workbook()
        ws = wb.active
        str_data = Article.articles_to_json(list_articles)
        data = json.loads(str_data)
        headers = list(data[0].keys()) if data else []
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)
        for row_index, row_data in enumerate(data, start=2):
            for col_index, header in enumerate(headers, start=1):
                ws.cell(row=row_index, column=col_index, value=row_data.get(header, ""))
        wb.save(excel_file_path)
        logger.info("Excel file created.")
        logger.info("Creating Output...")
        workitems.outputs.create(files=[excel_file_path])
